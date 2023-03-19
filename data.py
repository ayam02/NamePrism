#references: 
#https://pandas.pydata.org/docs/
#https://pandas.pydata.org/docs/reference/io.html#excel
#https://pandas.pydata.org/docs/user_guide/io.html#excel-files


from os import getcwd
from types import NoneType
import pandas as pd
import openpyxl
import excelcreator

#open the excel file
book = openpyxl.load_workbook('./CEO.xlsx')
sheet= book['Sheet1']

#read the excel file
content=pd.read_excel('./CEO.xlsx', sheet_name="Sheet1") 

def get_cols_num():
    global columns
    columns = sheet.max_column

def find_manualcolnum():
    for col in range(1,columns):
        if sheet.cell(1,col).value == 'Ethnicity of CEO (Asian, Black, Hispanic, Other, White) Picture/Name':
            return col

def find_nameprismcol():
    for col in range(1,columns):
        if sheet.cell(1,col).value == 'Ethnicity of CEO (Asian, Black, Hispanic, Other, White) NamePrism':
            return col
           

#find the number of matching data
def matching_data():
    global matching
    matching=0
    rowindex=0
    excelcreator.create_book(r"C:\Users\ayasm\OneDrive\Desktop\year3\OS\excel2.xlsx")
    get_rows_num() 
    get_cols_num()

    row=2
    for row in range(2,rows):
        matchingArray=[]
        if (type(sheet.cell(row,find_manualcolnum()).value)== NoneType or  type(sheet.cell(row,find_nameprismcol()).value)==NoneType):
            continue
        if(sheet.cell(row, manual_col).value.upper().__contains__(sheet.cell(row,nameprism_col).value.upper())):
            #print(sheet.cell(row, nameprism_col).value.upper(),"        ", sheet.cell(row,manual_col).value.upper())
            matching= matching+1
    return matching

def get_total():
    return (rows-1)

def book_close():
    book.close()

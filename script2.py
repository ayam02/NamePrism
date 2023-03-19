import openpyxl
import pandas as pd

#open the excel file
book = openpyxl.load_workbook(r'./excel2.xlsx')
sheet= book['Sheet']

#read the excel file
ethnic_to_nonEthnic=0
nonEthnic_to_ethnic = 0

col_1 =0
col_2 =0

print(sheet.max_row)

for column in range(1, sheet.max_column):
    if sheet.cell(1, column).value == 'Ethnicity of CEO (Asian, Black, Hispanic, Other, White) NamePrism':
        col_1 = column
    elif sheet.cell(1,column).value == 'Ethnicity of CEO (Asian, Black, Hispanic, Other, White) Picture/Name':
        col_2 = column

for row in range (1, sheet.max_row):
    if sheet.cell(row, col_1).value == 'WHITE':
        ethnic_to_nonEthnic = ethnic_to_nonEthnic+1
    else:
        nonEthnic_to_ethnic = nonEthnic_to_ethnic+1

print("non ethnic CEOs that were classified as ethnic are " , nonEthnic_to_ethnic)
print("ethnic CEOs that were classified as non ethnic are " , ethnic_to_nonEthnic)



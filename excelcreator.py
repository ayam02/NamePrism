import openpyxl 
import data

wb= openpyxl.Workbook()
sheet = wb.active
row1 = ["old_company", "full_name", "Ethnicity of CEO (Asian, Black, Hispanic, Other, White) NamePrism", "Ethnicity of CEO (Asian, Black, Hispanic, Other, White) Picture/Name"]

def initialize():
    for data in row1:
        sheet.cell(1, row1.index(data)+1).value= data

def add_data (data):
    x=0
    while(x < len(data)):
            row = sheet.max_row+1
            for i in range(1, len(row1)+1):
                #print('index= ',x)
                print(data[x])
                print('row= ',row," column = ", i)
                sheet.cell(row, i).value= data[x]
                x=x+1
            print("\n")
    wb.save( r"C:\Users\ayasm\OneDrive\Desktop\year3\OS\excel2.xlsx")




def create_book (path):
    wb.save(path)
    initialize()
    wb.save(path)

wb.close()

